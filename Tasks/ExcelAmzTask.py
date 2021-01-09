import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
class Sample:
    def read_data(self):
        driver = webdriver.Chrome(executable_path=r"C:\Users\KrishnaPriya\PycharmProjects\SeleniumWIthPython\Driver\chromedriver.exe")
        driver.get("http://www.amazon.in")
        driver.maximize_window()
        time.sleep(2)
        txt_search = driver.find_element(By.ID, "twotabsearchtextbox")
        txt_search.send_keys("iphone")
        searchicon = driver.find_element(By.XPATH, "//div[@class='nav-search-submit nav-sprite']")
        searchicon.click()
        time.sleep(10)
        proddetails = driver.find_elements(By.XPATH,"(//div[@class='sg-col-4-of-12 sg-col-8-of-16 sg-col-16-of-24 sg-col-12-of-20 sg-col-24-of-32 sg-col sg-col-28-of-36 sg-col-20-of-28'])")
        rowcount = len(proddetails)
        excel_loc = r"C:\Users\KrishnaPriya\Desktop\Selenium\AmzIphonedata.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet("ProductDetails", 0)
        #rows = sheet.max_row
        column = sheet.max_column
        #for i in proddetails:
        #print(p)
        for m in range(1,rowcount+1):
            for n in range(1,column+1):
                firstpath="(//div[@class='sg-col-4-of-12 sg-col-8-of-16 sg-col-16-of-24 sg-col-12-of-20 sg-col-24-of-32 sg-col sg-col-28-of-36 sg-col-20-of-28'])"
                secondpath="["
                thirdpath="]"
                finalpath=firstpath+secondpath+str(m)+thirdpath
                data=driver.find_element_by_xpath(finalpath)
                p=data.text
                sheet.cell(1,1).value="Iphone models"
                sheet.cell(m,2).value=p
                workbook.save(excel_loc)
s=Sample()
s.read_data()


