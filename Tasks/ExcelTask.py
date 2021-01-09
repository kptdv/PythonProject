import openpyxl
class ExcelSample:
    def read_file(self):
        xlsx_path=r"C:\Users\KrishnaPriya\Desktop\Selenium\SeleniumTestData.xlsx"
        file=openpyxl.load_workbook(xlsx_path)
        activesheet=file.active

        row=activesheet.max_row
        column=activesheet.max_column

        for i in range(1,row+1):
            for j in range(1,column+1):
                printvalue=activesheet.cell(i,j).value
                print(printvalue)
e=ExcelSample()
e.read_file()

